"""
excel_report.py
Auto-generates a professional multi-sheet Excel workbook with:
  - Executive KPI Summary
  - Product-level Optimization Table
  - ML Model Metrics
  - Embedded Charts (demand trend, cost breakdown, EOQ)
  - Conditional formatting and styled tables
"""

import os
import json
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import datetime

REPORTS_DIR = os.path.join(os.path.dirname(__file__), "..", "outputs", "reports")
CHARTS_DIR  = os.path.join(os.path.dirname(__file__), "..", "outputs", "charts")
os.makedirs(REPORTS_DIR, exist_ok=True)

# ─── Style constants ──────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
SUBHDR_FILL   = PatternFill("solid", fgColor="2E75B6")
ALT_ROW_FILL  = PatternFill("solid", fgColor="D6E4F7")
GOOD_FILL     = PatternFill("solid", fgColor="C6EFCE")
BAD_FILL      = PatternFill("solid", fgColor="FFC7CE")
WARN_FILL     = PatternFill("solid", fgColor="FFEB9C")
WHITE_FONT    = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
BOLD_FONT     = Font(bold=True, name="Calibri", size=11)
BODY_FONT     = Font(name="Calibri", size=10)
TITLE_FONT    = Font(bold=True, name="Calibri", size=16, color="1F4E79")
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _style_header_row(ws, row: int, n_cols: int):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = HEADER_FILL
        cell.font      = WHITE_FONT
        cell.alignment = CENTER
        cell.border    = THIN_BORDER


def _style_data_rows(ws, start_row: int, end_row: int, n_cols: int):
    for r in range(start_row, end_row + 1):
        fill = ALT_ROW_FILL if r % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill      = fill
            cell.font      = BODY_FONT
            cell.alignment = CENTER
            cell.border    = THIN_BORDER


def _autofit_columns(ws, min_width=10, max_width=30):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_width, max(min_width, max_len + 2))


# ─── Sheet 1: Executive Summary ───────────────────────────────────────────────

def write_executive_summary(wb, opt_df: pd.DataFrame):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False

    # Title banner
    ws.merge_cells("A1:H1")
    ws["A1"] = "📦  SUPPLY CHAIN ANALYTICS — EXECUTIVE SUMMARY"
    ws["A1"].font      = TITLE_FONT
    ws["A1"].alignment = CENTER
    ws["A1"].fill      = PatternFill("solid", fgColor="EBF3FB")

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%B %d, %Y  %H:%M')}  |  Data Period: 2022–2025"
    ws["A2"].font      = Font(name="Calibri", size=10, italic=True, color="555555")
    ws["A2"].alignment = CENTER

    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 18

    # KPI boxes — row 4
    kpis = [
        ("Total Revenue",        f"${opt_df['total_revenue'].sum():,.0f}",         "1E8449"),
        ("Total Products",       str(len(opt_df)),                                  "1F4E79"),
        ("Avg Fill Rate",        f"{opt_df['fill_rate_%'].mean():.1f}%",            "1E8449"),
        ("Avg Stockout Rate",    f"{opt_df['stockout_rate_%'].mean():.1f}%",        "C0392B"),
        ("Avg Inventory Turnover",f"{opt_df['inventory_turnover'].mean():.1f}x",   "7D3C98"),
        ("Total Supply Cost",    f"${opt_df['total_supply_cost'].sum():,.0f}",      "C0392B"),
        ("Avg On-Time Delivery", f"{opt_df['on_time_delivery_%'].mean():.1f}%",    "1E8449"),
        ("Avg Gross Margin",     f"{opt_df['gross_margin_%'].mean():.1f}%",        "1E8449"),
    ]

    for i, (label, value, color) in enumerate(kpis):
        col = i + 1
        ws.cell(row=4, column=col, value=label).font  = Font(bold=True, name="Calibri",
                                                              size=9, color=color)
        ws.cell(row=4, column=col).alignment = CENTER
        ws.cell(row=4, column=col).fill = PatternFill("solid", fgColor="F4F6F9")
        ws.cell(row=5, column=col, value=value).font  = Font(bold=True, name="Calibri",
                                                              size=16, color=color)
        ws.cell(row=5, column=col).alignment = CENTER
        ws.row_dimensions[5].height = 32

    # Category summary table — row 7
    ws.cell(row=7, column=1, value="Category Performance Summary").font = BOLD_FONT
    cat_summary = opt_df.groupby("category").agg(
        Products=("product_id","count"),
        Revenue=("total_revenue","sum"),
        Fill_Rate=("fill_rate_%","mean"),
        Stockout_Rate=("stockout_rate_%","mean"),
        Supply_Cost=("total_supply_cost","sum"),
    ).reset_index().round(2)
    cat_summary.columns = ["Category","# Products","Revenue ($)",
                            "Fill Rate (%)","Stockout Rate (%)","Supply Cost ($)"]

    headers = list(cat_summary.columns)
    for ci, h in enumerate(headers, 1):
        ws.cell(row=8, column=ci, value=h)
    _style_header_row(ws, 8, len(headers))
    for ri, row_data in enumerate(cat_summary.itertuples(index=False), 9):
        for ci, val in enumerate(row_data, 1):
            ws.cell(row=ri, column=ci, value=val)
    _style_data_rows(ws, 9, 8 + len(cat_summary), len(headers))
    _autofit_columns(ws)

    print("   ✅ Sheet: Executive Summary")


# ─── Sheet 2: Product Optimization ───────────────────────────────────────────

def write_product_optimization(wb, opt_df: pd.DataFrame):
    ws = wb.create_sheet("Product Optimization")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:N1")
    ws["A1"] = "Product-Level Optimization — EOQ, Safety Stock & KPIs"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    cols = ["product_id","category","unit_cost","eoq","opt_safety_stock",
            "opt_reorder_point","fill_rate_%","stockout_rate_%",
            "inventory_turnover","on_time_delivery_%",
            "total_revenue","total_supply_cost","gross_margin_%","avg_reliability"]
    display_cols = ["Product","Category","Unit Cost ($)","EOQ (units)",
                    "Safety Stock","Reorder Point","Fill Rate (%)","Stockout Rate (%)",
                    "Inv. Turnover","On-Time Del. (%)","Revenue ($)","Supply Cost ($)",
                    "Gross Margin (%)","Supplier Reliability"]

    sub = opt_df[cols].round(2)
    for ci, h in enumerate(display_cols, 1):
        ws.cell(row=2, column=ci, value=h)
    _style_header_row(ws, 2, len(display_cols))

    for ri, row_data in enumerate(sub.itertuples(index=False), 3):
        for ci, val in enumerate(row_data, 1):
            ws.cell(row=ri, column=ci, value=round(float(val), 2)
                     if isinstance(val, (int, float, np.floating)) else val)

    _style_data_rows(ws, 3, 2 + len(sub), len(display_cols))

    # Conditional formatting via cell colours
    fill_rate_col   = display_cols.index("Fill Rate (%)") + 1
    stockout_col    = display_cols.index("Stockout Rate (%)") + 1
    margin_col      = display_cols.index("Gross Margin (%)") + 1

    for ri in range(3, 3 + len(sub)):
        fr_cell = ws.cell(row=ri, column=fill_rate_col)
        if fr_cell.value and fr_cell.value >= 95:
            fr_cell.fill = GOOD_FILL
        elif fr_cell.value and fr_cell.value < 85:
            fr_cell.fill = BAD_FILL
        else:
            fr_cell.fill = WARN_FILL

        so_cell = ws.cell(row=ri, column=stockout_col)
        if so_cell.value and so_cell.value <= 5:
            so_cell.fill = GOOD_FILL
        elif so_cell.value and so_cell.value > 15:
            so_cell.fill = BAD_FILL
        else:
            so_cell.fill = WARN_FILL

        mg_cell = ws.cell(row=ri, column=margin_col)
        if mg_cell.value and mg_cell.value >= 60:
            mg_cell.fill = GOOD_FILL
        elif mg_cell.value and mg_cell.value < 30:
            mg_cell.fill = BAD_FILL

    _autofit_columns(ws)
    print("   ✅ Sheet: Product Optimization")


# ─── Sheet 3: ML Model Metrics ────────────────────────────────────────────────

def write_ml_metrics(wb, metrics_path: str):
    ws = wb.create_sheet("ML Model Metrics")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"] = "Machine Learning Model Performance"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    try:
        with open(metrics_path) as f:
            metrics = json.load(f)
    except FileNotFoundError:
        metrics = {"random_forest": {"MAE": "N/A", "RMSE": "N/A", "MAPE": "N/A", "R2": "N/A"},
                   "xgboost":       {"MAE": "N/A", "RMSE": "N/A", "MAPE": "N/A", "R2": "N/A"}}

    headers = ["Model", "MAE", "RMSE", "MAPE (%)", "R² Score", "Interpretation"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=2, column=ci, value=h)
    _style_header_row(ws, 2, len(headers))

    interpret = {
        "random_forest": "Ensemble of decision trees; robust to noise",
        "xgboost":       "Gradient boosted trees; high accuracy & speed",
    }

    for ri, (model_name, m) in enumerate(metrics.items(), 3):
        r2  = m.get("R2", "N/A")
        row = [model_name.replace("_", " ").title(),
               m.get("MAE"), m.get("RMSE"), m.get("MAPE"), r2,
               interpret.get(model_name, "")]
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
        ws.cell(row=ri, column=5).fill = (
            GOOD_FILL if isinstance(r2, float) and r2 > 0.85 else WARN_FILL)

    _style_data_rows(ws, 3, 4, len(headers))
    _autofit_columns(ws)

    # Notes section
    ws.cell(row=7, column=1, value="Metric Definitions").font = BOLD_FONT
    defs = [
        ("MAE",       "Mean Absolute Error — average magnitude of prediction errors (units)"),
        ("RMSE",      "Root Mean Square Error — penalises large errors more heavily"),
        ("MAPE (%)",  "Mean Absolute Percentage Error — relative forecast accuracy"),
        ("R² Score",  "Coefficient of determination — 1.0 = perfect fit"),
    ]
    for i, (metric, definition) in enumerate(defs, 8):
        ws.cell(row=i, column=1, value=metric).font = BOLD_FONT
        ws.cell(row=i, column=2, value=definition).font = BODY_FONT
        ws.merge_cells(f"B{i}:F{i}")

    print("   ✅ Sheet: ML Model Metrics")


# ─── Sheet 4: Embedded Charts ─────────────────────────────────────────────────

def write_charts_sheet(wb, opt_df: pd.DataFrame):
    ws = wb.create_sheet("Charts & Visuals")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:P1")
    ws["A1"] = "Supply Chain Analytics — Key Visualisations"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    # Embed saved PNG charts
    chart_files = [
        ("demand_trend.png",        "A3",  "Monthly Demand Trend"),
        ("forecast_vs_actual.png",  "A32", "ML Forecast vs Actual"),
        ("anomaly_detection.png",   "A61", "Anomaly Detection"),
        ("eoq_comparison.png",      "M3",  "EOQ by Product"),
        ("cost_breakdown.png",      "M32", "Cost Breakdown"),
        ("stockout_vs_fillrate.png","M61", "Stockout vs Fill Rate"),
    ]

    for fname, anchor, title in chart_files:
        fpath = os.path.join(CHARTS_DIR, fname)
        if os.path.exists(fpath):
            img = XLImage(fpath)
            img.width  = 480
            img.height = 220
            ws.add_image(img, anchor)
        else:
            ws[anchor] = f"[Chart not found: {title}]"

    # Built-in openpyxl bar chart — Revenue by Category
    cat_data = opt_df.groupby("category")["total_revenue"].sum().reset_index()
    data_start = 90
    ws.cell(row=data_start, column=1, value="Category")
    ws.cell(row=data_start, column=2, value="Revenue")
    for i, (_, row) in enumerate(cat_data.iterrows(), 1):
        ws.cell(row=data_start + i, column=1, value=row["category"])
        ws.cell(row=data_start + i, column=2, value=round(row["total_revenue"], 0))

    chart = BarChart()
    chart.type  = "col"
    chart.title = "Revenue by Category"
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Category"
    data_ref  = Reference(ws, min_col=2, min_row=data_start,
                           max_row=data_start + len(cat_data))
    cats_ref  = Reference(ws, min_col=1, min_row=data_start + 1,
                           max_row=data_start + len(cat_data))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape   = 4
    chart.width   = 18
    chart.height  = 12
    ws.add_chart(chart, "A90")

    print("   ✅ Sheet: Charts & Visuals")


# ─── Main Builder ─────────────────────────────────────────────────────────────

def build_excel_report(opt_df: pd.DataFrame,
                        metrics_path: str = None,
                        output_filename: str = "supply_chain_analytics_report.xlsx"):
    if metrics_path is None:
        metrics_path = os.path.join(REPORTS_DIR, "model_metrics.json")

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    write_executive_summary(wb, opt_df)
    write_product_optimization(wb, opt_df)
    write_ml_metrics(wb, metrics_path)
    write_charts_sheet(wb, opt_df)

    out_path = os.path.join(REPORTS_DIR, output_filename)
    wb.save(out_path)
    print(f"\n✅ Excel report saved: {out_path}")
    return out_path


if __name__ == "__main__":
    from data_preprocessing import load_data, clean_data, engineer_features, encode_categoricals
    from optimization import generate_optimization_report

    df  = encode_categoricals(engineer_features(clean_data(load_data())))
    opt = generate_optimization_report(df)
    build_excel_report(opt)
